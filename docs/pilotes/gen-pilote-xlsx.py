"""
Génère le fichier Excel à distribuer aux testeurs pilotes Maïa.
3 onglets : Lisez-moi / Public+Élève / Prof
Chaque ligne URL a 3 colonnes vides (Réponse / Notes / Remarques) avec
hauteur ~75px pour permettre la prise de notes.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ──────────────────────────────────────────────────────────────

FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", size=14, bold=True)
FONT_INTRO = Font(name="Arial", size=10, italic=True, color="666666")
FONT_SCENARIO = Font(name="Arial", size=10, bold=True, color="4338CA")

FILL_HEADER = PatternFill("solid", start_color="4338CA")  # indigo-700
FILL_ALT = PatternFill("solid", start_color="F1F5F9")  # slate-100
FILL_SCENARIO = PatternFill("solid", start_color="EEF2FF")  # indigo-50

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

ROW_HEIGHT = 75  # pixels, ~3 lignes de texte


def style_header(cell):
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER


def style_body(cell, alt=False, font=None):
    cell.font = font or FONT_BODY
    cell.alignment = ALIGN_LEFT_TOP
    cell.border = THIN_BORDER
    if alt:
        cell.fill = FILL_ALT


def style_scenario(cell):
    cell.font = FONT_SCENARIO
    cell.fill = FILL_SCENARIO
    cell.alignment = ALIGN_LEFT_CENTER
    cell.border = THIN_BORDER


# ── Données ─────────────────────────────────────────────────────────────

PUBLIC_TESTS = [
    ("Landing", "/", "Comprendre la promesse Maïa, fluide ?"),
    ("Pilotes", "/pilotes", "Page candidature école — CTA mailto s'ouvre ?"),
    ("RGPD simple", "/legal/lecture-facile", "Texte simplifié à montrer aux parents"),
    ("RGPD complet", "/legal/confidentialite", "Politique confidentialité complète"),
    ("CGU", "/legal/cgu", "Conditions d'utilisation"),
    ("Cookies", "/legal/cookies", "Politique cookies"),
    ("Mentions", "/legal/mentions-legales", "Mentions légales BE"),
    ("404 brandée", "/route-inexistante", "Page 404 Maïa s'affiche ?"),
    ("Login", "/login", "Boutons SSO Google + Microsoft (bientôt)"),
]

ELEVE_TESTS = [
    # Onboarding
    ("Onboarding", "/onboarding/name", "Saisir prénom + nom"),
    ("Onboarding", "/onboarding/consent-rgpd", "Cocher consentement (ou parent si <16 ans)"),
    ("Onboarding", "/onboarding/pin-setup", "Créer PIN 4 chiffres"),
    ("Onboarding", "/onboarding/join-class", "Saisir code classe 8 chars du prof"),
    # Re-auth
    ("Re-auth quotidienne", "/auth/pin-unlock", "PIN demandé chaque matin, 3 essais"),
    # Accueil
    ("Accueil", "/accueil", "Heatmap multi-matière + card Plan Maïa"),
    ("Drill-down concept", "/accueil/concepts/[id]", "Click cellule heatmap → théorie + pièges"),
    # Devoirs
    ("Devoirs", "/accueil/devoirs", "Liste 'À faire / Récents'"),
    ("Devoir détail", "/accueil/devoirs/[id]", "Brief + bouton Démarrer le quiz"),
    ("Quiz devoir", "/accueil/devoirs/[id]/quiz", "Répondre, indices tuteur, solution, 👍/👎"),
    ("Bilan post-quiz", "/accueil/devoirs/[id]/bilan", "Score + KPIs + mastery par concept"),
    # Plan Maïa
    ("Plan Maïa overview", "/accueil/plan-maia/today", "Liste questions du jour + raisons"),
    ("Quiz Plan Maïa", "/accueil/plan-maia/today/quiz", "Pick-and-choose, skip n'importe quand"),
    ("Bilan Plan Maïa", "/accueil/plan-maia/today/bilan", "Auto-affiché après dernière question"),
    # Live
    ("Live (Kahoot mode)", "/accueil/live/[code]", "Rejoindre session live via code prof"),
    # Settings
    ("Paramètres", "/accueil/parametres/compte", "Voir compte, classes, switch OpenDyslexic"),
    ("Confidentialité", "/accueil/parametres/confidentialite", "Révoquer/donner consents"),
    ("Export RGPD", "/accueil/parametres/export-donnees", "Demander export Art.20"),
    ("Suppression compte", "/accueil/parametres/suppression-compte", "Anonymisation (en fin de test)"),
]

PROF_TESTS = [
    # Onboarding
    ("Onboarding", "/onboarding/name", "Prénom + nom"),
    ("Onboarding", "/onboarding/teaching-levels", "Niveaux enseignés (5e à terminale)"),
    ("Onboarding", "/onboarding/pin-setup", "PIN 4 chiffres"),
    # Accueil
    ("Accueil prof", "/accueil", "Espace enseignant + greeting"),
    # Classes
    ("Classes", "/accueil/classes", "Liste classes"),
    ("Créer classe", "/accueil/classes/nouvelle", "Créer une nouvelle classe"),
    ("Classe détail", "/accueil/classes/[id]", "Membres, code invitation 8 chars, devoirs"),
    ("Invitation classe", "/accueil/classes/[id]/invitation", "QR code + lien partageable"),
    # Cours & Ingestion
    ("Cours", "/accueil/cours", "Liste cours importés"),
    ("Import PDF", "/accueil/import", "Importer un syllabus PDF"),
    ("Ingestion job", "/accueil/ingestion/[jobId]", "Suivi job ingestion live"),
    ("Exercices cours", "/accueil/cours/[id]/exercices", "Exercices générés par Maïa"),
    ("Détail exo", "/accueil/cours/[id]/exercices/[id]", "Exo + KaTeX MathML"),
    # Curation
    ("Curation liste", "/accueil/curation", "Questions à valider / rejeter"),
    ("Curation concept", "/accueil/curation/concept/[id]", "Théorie + misconceptions + HINTS CRUD"),
    # Devoirs prof
    ("Créer devoir", "/accueil/classes/[id]/devoirs/nouveau", "Cours + sélection questions"),
    ("Devoir vue d'ensemble", "/accueil/classes/[id]/devoirs/[id]", "Élèves + Heatmap + Top erreurs"),
    ("Drill-down élève", "/accueil/classes/[id]/devoirs/[id]/eleve/[id]", "Click cellule heatmap → détail"),
    # Stats
    ("Stats direction", "/accueil/stats-direction", "KPI agrégés + Top 5 concepts faibles"),
    # Live
    ("Créer session live", "/accueil/session/nouvelle", "Sélectionner questions live"),
    ("Live lobby", "/accueil/live", "Sessions actives"),
    ("Manager live", "/accueil/live/[id]", "Realtime réponses + random-pick projeté"),
    # Misc
    ("Horaire", "/accueil/horaire", "Grille hebdomadaire cours"),
    ("Paramètres", "/accueil/parametres/compte", "Compte + OpenDyslexic"),
]

COLS = ["#", "Catégorie", "URL", "Quoi tester", "Réponse", "Notes", "Remarques"]
COL_WIDTHS = [5, 22, 50, 50, 28, 28, 28]


def build_sheet(ws, title, intro_lines, tests):
    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Intro rows
    current_row = 2
    for line in intro_lines:
        ws.merge_cells(f"A{current_row}:G{current_row}")
        ws[f"A{current_row}"] = line
        ws[f"A{current_row}"].font = FONT_INTRO
        ws[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # Header row
    current_row += 1
    for col_idx, col_name in enumerate(COLS, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        style_header(cell)
    ws.row_dimensions[current_row].height = 25

    # Test rows
    for i, (categorie, url, description) in enumerate(tests, start=1):
        current_row += 1
        ws.row_dimensions[current_row].height = ROW_HEIGHT

        # # (numéro)
        c = ws.cell(row=current_row, column=1, value=i)
        c.font = FONT_BODY
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

        # Catégorie
        c = ws.cell(row=current_row, column=2, value=categorie)
        style_scenario(c)

        # URL (en monospace italic conceptuel)
        c = ws.cell(row=current_row, column=3, value=url)
        c.font = Font(name="Courier New", size=10)
        c.alignment = ALIGN_LEFT_TOP
        c.border = THIN_BORDER

        # Description
        c = ws.cell(row=current_row, column=4, value=description)
        style_body(c)

        # Réponse / Notes / Remarques (vides pour le testeur)
        for col in [5, 6, 7]:
            c = ws.cell(row=current_row, column=col, value="")
            c.border = THIN_BORDER
            c.alignment = ALIGN_LEFT_TOP
            c.fill = PatternFill("solid", start_color="FFFFFF")

    # Column widths
    for idx, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Freeze header
    ws.freeze_panes = ws.cell(row=current_row - len(tests) + 1, column=1)


def build_lisez_moi(ws):
    ws["A1"] = "Guide de test pilote — Maïa MVP"
    ws["A1"].font = Font(name="Arial", size=18, bold=True, color="4338CA")
    ws.row_dimensions[1].height = 32

    sections = [
        ("À qui s'adresse ce document",
         "Aux 3 écoles pilotes Maïa (rentrée 2026). Profs et élèves désignés référents pour tester l'application avant ouverture officielle."),
        ("Comment l'utiliser",
         "1. Ouvre l'onglet correspondant à ton rôle (Public, Élève ou Prof)\n"
         "2. Pour chaque URL, ouvre-la dans ton navigateur connecté à Maïa\n"
         "3. Teste ce qui est demandé dans la colonne 'Quoi tester'\n"
         "4. Remplis les colonnes 'Réponse' (Marche / Bug / Pas testé), 'Notes' (détails) et 'Remarques' (suggestions)"),
        ("Base URL",
         "https://maia.app (ou l'URL preview Vercel qui t'a été communiquée). Préfixe toutes les URLs relatives par cette base."),
        ("URLs avec [id] / [code]",
         "Certaines URLs contiennent un [id] qui change selon ta classe / ton compte. Tu y accèdes en cliquant les liens dans l'app (pas en tapant l'URL à la main). L'URL est listée pour référence."),
        ("Comment remonter un bug",
         "Envoie un mail à pilotes@maia.app avec :\n"
         "  • URL exacte\n"
         "  • Rôle (élève / prof)\n"
         "  • Sévérité : 🔴 Bloquant / 🟡 UX gênant / 🟢 Cosmétique\n"
         "  • Ce qui était attendu vs ce qui s'est passé\n"
         "  • Capture d'écran si possible"),
        ("Nouveautés Sprint 5-6 (à observer particulièrement)",
         "🆕 Drill-down élève × concept depuis heatmap prof\n"
         "🆕 CRUD des indices (hints) du tuteur Maïa par le prof\n"
         "🆕 Bilan post-quiz détaillé (page persistante)\n"
         "🆕 Plan Maïa quotidien quiz pick-and-choose dédié\n"
         "🆕 Switch OpenDyslexic dans Paramètres > Mon compte\n"
         "🆕 Lecture facile RGPD (texte simplifié pour parents)\n"
         "🆕 Stats direction agrégées toutes classes du prof\n"
         "🆕 Cron nuit pré-génère le Plan Maïa avant 1h Brussels"),
        ("Contact",
         "pilotes@maia.app — réponse sous 7 jours ouvrés (souvent plus vite)"),
    ]

    row = 3
    for title, body in sections:
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Arial", size=12, bold=True, color="4338CA")
        ws.row_dimensions[row].height = 22
        row += 1

        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = body
        ws[f"A{row}"].font = FONT_BODY
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        n_lines = body.count("\n") + 1
        ws.row_dimensions[row].height = max(40, 18 * n_lines + 8)
        row += 2

    ws.column_dimensions["A"].width = 100


def main(out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lisez-moi"
    build_lisez_moi(ws)

    ws_public = wb.create_sheet("Public")
    build_sheet(
        ws_public,
        "Tests pages publiques (sans connexion)",
        [
            "Ces URLs sont accessibles sans login. Vérifier la qualité avant de connecter ton compte.",
            "Préfixe : https://maia.app",
        ],
        PUBLIC_TESTS,
    )

    ws_eleve = wb.create_sheet("Élève")
    build_sheet(
        ws_eleve,
        "Tests parcours élève (compte étudiant)",
        [
            "Connecte-toi avec un compte élève via /login (Google SSO).",
            "URLs avec [id] : accède via les liens internes (pas en tapant l'URL à la main).",
            "Préfixe : https://maia.app",
        ],
        ELEVE_TESTS,
    )

    ws_prof = wb.create_sheet("Prof")
    build_sheet(
        ws_prof,
        "Tests parcours prof (compte enseignant)",
        [
            "Connecte-toi avec un compte prof. Tu dois avoir créé au moins 1 classe + importé 1 cours pour tester l'ensemble du flow.",
            "URLs avec [id] : accède via les liens internes.",
            "Préfixe : https://maia.app",
        ],
        PROF_TESTS,
    )

    wb.save(out_path)
    print(f"OK saved: {out_path}")
    print(f"  Lisez-moi : 1 sheet")
    print(f"  Public : {len(PUBLIC_TESTS)} URLs")
    print(f"  Élève : {len(ELEVE_TESTS)} URLs")
    print(f"  Prof : {len(PROF_TESTS)} URLs")
    print(f"  Total : {len(PUBLIC_TESTS) + len(ELEVE_TESTS) + len(PROF_TESTS)} URLs")


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "Maia-Pilote-Test-Guide.xlsx"
    main(out)
